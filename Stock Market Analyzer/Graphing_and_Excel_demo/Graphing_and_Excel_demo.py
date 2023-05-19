import datetime
from datetime import timedelta, date
from Bond import Bond
from BondETF import BondETF
from Asset import Asset
from Portfolio import Portfolio
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.chart.axis import DateAxis
import numpy as np
import bisect
import math

'''
#Individual Asset BackTesting
cols = [0,1]
sData = pd.read_excel('./S&P500.xlsx',sheet_name=0,usecols=[0,4])
syData = pd.read_excel('./S&P500.xlsx',sheet_name=1,usecols=cols)
stockData = [(sData.Date[i].date(),sData.Close[i]) for i in range(len(sData.Date))]
stockYieldData = [(syData.Date[i].date(),syData.Value[i]) for i in range(683, len(syData.Date)) if (syData.Date[i].date().month)%3 == 0]
print("stockData length: " + str(len(stockData)) + " First date: " + stockData[0][0].strftime("%m/%d/%y"))
print("stockYieldData length: " + str(len(stockYieldData)) + " First date: " + stockYieldData[0][0].strftime("%m/%d/%y"))

stocks = Asset("Stocks", stockData, stockYieldData, freq=4)
assets = [stocks]
weights = [1.0]
permanentPortfolio = Portfolio("Permanent Portfolio", assets, weights, 17.66)
ppData = permanentPortfolio.generateData()
incomeData = permanentPortfolio.getIncomeData()
annualReturnsData = permanentPortfolio.getAnnualReturnData()

fileName = './S&P500.xlsx'
sheetName = 'Stock_Total_Return'
sheetName2 = 'Income Data'
sheetName3 = 'Annual Returns'

book = load_workbook(fileName)
writer = pd.ExcelWriter(fileName, engine = 'openpyxl')
df = pd.DataFrame(ppData, columns=['Date','Stocks', 'Value'])
df2 = pd.DataFrame(incomeData, columns = ['Date', 'Dividends'])
df3 = pd.DataFrame(annualReturnsData, columns = ['Date', 'Returns'])
try:
    book.remove(book[sheetName])
    book.remove(book[sheetName2])
    book.remove(book[sheetName3])
except:
    print("worksheet doesn't exist")
finally:
    writer.book = book
    df.to_excel(writer, sheet_name = sheetName, index=False)
    df2.to_excel(writer, sheet_name = sheetName2, index=False)
    df3.to_excel(writer, sheet_name = sheetName3, index=False)

wb = writer.book
sheet = wb[sheetName]
sheet2 = wb[sheetName2]
sheet3 = wb[sheetName3]
chart = LineChart()  
chart.height = 12
chart.width = 22
chart.title = " S&P 500 Total Return Value "
chart.x_axis.title = " Date "
chart.y_axis.title = " Value "
chart.y_axis.crossAx = 500
chart.x_axis = DateAxis(crossAx=100)
chart.x_axis.number_format = 'mm/dd/yyyy'
chart.x_axis.majorTimeUnit = "days"
values1 = Reference(sheet, min_col = 2, min_row = 2, max_row = len(ppData)+1) 
series1 = Series(values1, title="S&P500 Value")
chart.append(series1)
dates = Reference(sheet, min_col=1, min_row=2, max_row = len(ppData)+1)
chart.set_categories(dates) 
sheet.add_chart(chart, "G2") 

chart2 = BarChart()  
chart2.height = 12
chart2.width = 22
chart2.title = " Income Data "
chart2.x_axis.title = " Date "
chart2.y_axis.title = " amount "
chart2.y_axis.crossAx = 500
chart2.x_axis = DateAxis(crossAx=100)
chart2.x_axis.number_format = 'mm/dd/yyyy'
chart2.x_axis.majorTimeUnit = "months"
values2 = Reference(sheet2, min_col = 2, min_row = 2, max_row = len(incomeData)+1) 
series2 = Series(values2, title="Dividends")
chart2.append(series2)
dates2 = Reference(sheet2, min_col=1, min_row=2, max_row = len(incomeData)+1)
chart2.set_categories(dates2) 
sheet.add_chart(chart2, "G26") 
#sheet2.add_chart(chart2, "E2")

chart3 = BarChart()  
chart3.height = 12
chart3.width = 22
chart3.title = " Annual Returns "
chart3.x_axis.title = " Date "
chart3.y_axis.title = " Percent "
chart3.y_axis.crossAx = 500
chart3.x_axis = DateAxis(crossAx=100)
chart3.x_axis.number_format = 'mm/dd/yyyy'
chart3.x_axis.majorTimeUnit = "years"
values3 = Reference(sheet3, min_col = 2, min_row = 2, max_row = len(annualReturnsData)+1) 
series3 = Series(values3, title="Annual Returns")
chart3.append(series3)
dates3 = Reference(sheet3, min_col=1, min_row=2, max_row = len(annualReturnsData)+1)
chart3.set_categories(dates3) 
sheet.add_chart(chart3, "G50") 
#sheet3.add_chart(chart2, "E2")

writer.save()
writer.close()
'''
'''
#Stock and Bond

cols = [0,1]
sData = pd.read_excel('./S&P500.xlsx',sheet_name='S&P500' ,usecols=[0,4])
syData = pd.read_excel('./S&P500.xlsx',sheet_name='S&P Yield' ,usecols=cols)
bData = pd.read_excel('./tlt.xlsx', sheet_name='TLT Price and Yield (1970)', usecols=[0,1,2])
stockData = [(sData.Date[i].date(),sData.Close[i]) for i in range(10499, 23305) if (sData.Date[i+1].date().month - sData.Date[i].date().month)%12 == 1]
stockYieldData = [(syData.Date[i].date(),syData.Value[i]) for i in range(1188, len(syData.Date)-1) if (syData.Date[i].date().month)%3 == 0]
bondData = [(bData.Date[i].date(),bData.Value[i]) for i in range(len(bData.Date))]
bondYieldData = [(bData.Date[i].date(),bData.Yield[i]) for i in range(1, len(bData.Date))]
print("stockData length: " + str(len(stockData)) + " First date: " + stockData[0][0].strftime("%m/%d/%y"))
print("bondData length: " + str(len(bondData)) + " First date: " + bondData[0][0].strftime("%m/%d/%y"))
print("stockYieldData length: " + str(len(stockYieldData)) + " First date: " + stockYieldData[0][0].strftime("%m/%d/%y"))
print("bondYieldData length: " + str(len(bondYieldData)) + " First date: " + bondYieldData[0][0].strftime("%m/%d/%y"))

stocks = Asset("Stocks", stockData, stockYieldData, freq=4)
bonds = Asset("Bond", bondData, yData=bondYieldData, freq=12)
assets = [stocks, bonds]
weights = [0.71, 0.29]
permanentPortfolio = Portfolio("Stock and Bond", assets, weights, 100000)
ppData = permanentPortfolio.generateData()
incomeData = permanentPortfolio.getIncomeData()
annualReturnsData = permanentPortfolio.getAnnualReturnData()

fileName = './Stock and Bond.xlsx'
sheetName = 'SB Data (1970) Trial 2'
sheetName2 = 'Income Data (1970) Trial 2'
sheetName3 = 'Annual Returns (1970) Trial 2'

book = load_workbook(fileName)
writer = pd.ExcelWriter(fileName, engine = 'openpyxl')
df = pd.DataFrame(ppData, columns=['Date','Stocks', 'Bonds', 'Value'])
df2 = pd.DataFrame(incomeData, columns = ['Date', 'Dividends'])
df3 = pd.DataFrame(annualReturnsData, columns = ['Date', 'Returns'])
try:
    book.remove(book[sheetName])
    book.remove(book[sheetName2])
    book.remove(book[sheetName3])
except:
    print("worksheet doesn't exist")
finally:
    writer.book = book
    df.to_excel(writer, sheet_name = sheetName, index=False)
    df2.to_excel(writer, sheet_name = sheetName2, index=False)
    df3.to_excel(writer, sheet_name = sheetName3, index=False)

wb = writer.book
sheet = wb[sheetName]
sheet2 = wb[sheetName2]
sheet3 = wb[sheetName3]
chart = LineChart()  
chart.height = 12
chart.width = 22
chart.title = " S/B Account Value "
chart.x_axis.title = " Date "
chart.y_axis.title = " Value "
chart.y_axis.crossAx = 500
chart.x_axis = DateAxis(crossAx=100)
chart.x_axis.number_format = 'mm/dd/yyyy'
chart.x_axis.majorTimeUnit = "months"
values1 = Reference(sheet, min_col = 4, min_row = 2, max_row = len(ppData)+1) 
series1 = Series(values1, title="S/B Portfolio")
chart.append(series1)
values2 = Reference(sheet, min_col = 2, min_row = 2, max_row = len(ppData)+1) 
series2 = Series(values2, title="Stock Position")
chart.append(series2)
values3 = Reference(sheet, min_col = 3, min_row = 2, max_row = len(ppData)+1) 
series3 = Series(values3, title="Bond Position")
chart.append(series3)
dates = Reference(sheet, min_col=1, min_row=2, max_row = len(ppData)+1)
chart.set_categories(dates) 
sheet.add_chart(chart, "G2") 

chart2 = BarChart()  
chart2.height = 12
chart2.width = 22
chart2.title = " Income Data "
chart2.x_axis.title = " Date "
chart2.y_axis.title = " amount "
chart2.y_axis.crossAx = 500
chart2.x_axis = DateAxis(crossAx=100)
chart2.x_axis.number_format = 'mm/dd/yyyy'
chart2.x_axis.majorTimeUnit = "months"
values2 = Reference(sheet2, min_col = 2, min_row = 2, max_row = len(incomeData)+1) 
series2 = Series(values2, title="Dividends")
chart2.append(series2)
dates2 = Reference(sheet2, min_col=1, min_row=2, max_row = len(incomeData)+1)
chart2.set_categories(dates2) 
sheet.add_chart(chart2, "G26") 

chart3 = BarChart()  
chart3.height = 12
chart3.width = 22
chart3.title = " Annual Returns "
chart3.x_axis.title = " Date "
chart3.y_axis.title = " Percent "
chart3.y_axis.crossAx = 500
chart3.x_axis = DateAxis(crossAx=100)
chart3.x_axis.number_format = 'mm/dd/yyyy'
chart3.x_axis.majorTimeUnit = "years"
values3 = Reference(sheet3, min_col = 2, min_row = 2, max_row = len(annualReturnsData)+1) 
series3 = Series(values3, title="Annual Returns")
chart3.append(series3)
dates3 = Reference(sheet3, min_col=1, min_row=2, max_row = len(annualReturnsData)+1)
chart3.set_categories(dates3) 
sheet.add_chart(chart3, "G50") 

writer.save()
writer.close()
'''

#Permanent Portfolio Backtesting

cols = [0,1]
sData = pd.read_excel('./S&P500.xlsx',sheet_name='S&P500' ,usecols=[0,4])
syData = pd.read_excel('./S&P500.xlsx',sheet_name='S&P Yield' ,usecols=cols)
bData = pd.read_excel('./tlt.xlsx', sheet_name='TLT Price and Yield (1970)', usecols=[0,1,2])
gData = pd.read_excel('./Gold-Prices.xlsx',sheet_name='Monthly(1969)' ,usecols=cols)
stockData = [(sData.Date[i].date(),sData.Close[i]) for i in range(10499, 23305) if (sData.Date[i+1].date().month - sData.Date[i].date().month)%12 == 1]
stockYieldData = [(syData.Date[i].date(),syData.Value[i]) for i in range(1188, len(syData.Date)-1) if (syData.Date[i].date().month)%3 == 0]
bondData = [(bData.Date[i].date(),bData.Value[i]) for i in range(len(bData.Date))]
bondYieldData = [(bData.Date[i].date(),bData.Yield[i]) for i in range(1, len(bData.Date))]
goldData =  [(gData.Date[i].date(),gData.Value[i]) for i in range(1, len(gData.Date)-1)]
print("stockData length: " + str(len(stockData)) + " First date: " + stockData[0][0].strftime("%m/%d/%y"))
print("bondData length: " + str(len(bondData)) + " First date: " + bondData[0][0].strftime("%m/%d/%y"))
print("goldData length: " + str(len(goldData)) + " First date: " + goldData[0][0].strftime("%m/%d/%y"))
print("stockYieldData length: " + str(len(stockYieldData)) + " First date: " + stockYieldData[0][0].strftime("%m/%d/%y"))
print("bondYieldData length: " + str(len(bondYieldData)) + " First date: " + bondYieldData[0][0].strftime("%m/%d/%y"))
#for i in range(len(bondYieldData)):
    #print("Bond Yield  " + bondYieldData[i][0].strftime("%m/%d/%y") + "   " + str(bondYieldData[i][1]))
#for i in range(len(stockYieldData)):
    #print("Stock Yield  " + stockYieldData[i][0].strftime("%m/%d/%y") + "   " + str(stockYieldData[i][1]))

stocks = Asset("Stocks", stockData, stockYieldData, freq=4)
bonds = Asset("Bond", bondData, yData=bondYieldData, freq=12)
gold = Asset("Gold", goldData)
assets = [stocks, bonds, gold]
weights = [0.334, 0.333, 0.333]
yearlyInflows = 50000
endDate = datetime.date(1990, 1, 20)
monthlyInflows = yearlyInflows/12.
permanentPortfolio = Portfolio("Permanent Portfolio", assets, weights, 30000)
ppData = permanentPortfolio.generateData(monthlyInflows, endDate = endDate)
incomeData = permanentPortfolio.getIncomeData()
annualReturnsData = permanentPortfolio.getAnnualReturnData()

fileName = './Permanent Portfolio.xlsx'
sheetName = 'PP w Inflows (1970)'
sheetName2 = 'Income Data w Inflows (1970)'
sheetName3 = 'Annual Returns w Inflows (1970)'

book = load_workbook(fileName)
writer = pd.ExcelWriter(fileName, engine = 'openpyxl')
df = pd.DataFrame(ppData, columns=['Date','Stocks', 'Bonds', 'Gold', 'Value'])
df2 = pd.DataFrame(incomeData, columns = ['Date', 'Dividends'])
df3 = pd.DataFrame(annualReturnsData, columns = ['Date', 'Returns'])
try:
    book.remove(book[sheetName])
    book.remove(book[sheetName2])
    book.remove(book[sheetName3])
except:
    print("worksheet doesn't exist")
finally:
    writer.book = book
    df.to_excel(writer, sheet_name = sheetName, index=False)
    df2.to_excel(writer, sheet_name = sheetName2, index=False)
    df3.to_excel(writer, sheet_name = sheetName3, index=False)

wb = writer.book
sheet = wb[sheetName]
sheet2 = wb[sheetName2]
sheet3 = wb[sheetName3]
chart = LineChart()  
chart.height = 12
chart.width = 22
chart.title = " PP Account Value "
chart.x_axis.title = " Date "
chart.y_axis.title = " Value "
chart.y_axis.crossAx = 500
chart.x_axis = DateAxis(crossAx=100)
chart.x_axis.number_format = 'mm/dd/yyyy'
chart.x_axis.majorTimeUnit = "months"
values1 = Reference(sheet, min_col = 5, min_row = 2, max_row = len(ppData)+1) 
series1 = Series(values1, title="Permanent Portfolio")
chart.append(series1)
values2 = Reference(sheet, min_col = 2, min_row = 2, max_row = len(ppData)+1) 
series2 = Series(values2, title="Stock Position")
chart.append(series2)
values3 = Reference(sheet, min_col = 3, min_row = 2, max_row = len(ppData)+1) 
series3 = Series(values3, title="Bond Position")
chart.append(series3)
values4 = Reference(sheet, min_col = 4, min_row = 2, max_row = len(ppData)+1) 
series4 = Series(values4, title="Gold Position")
chart.append(series4)
dates = Reference(sheet, min_col=1, min_row=2, max_row = len(ppData)+1)
chart.set_categories(dates) 
sheet.add_chart(chart, "G2") 

chart2 = BarChart()  
chart2.height = 12
chart2.width = 22
chart2.title = " Income Data "
chart2.x_axis.title = " Date "
chart2.y_axis.title = " amount "
chart2.y_axis.crossAx = 500
chart2.x_axis = DateAxis(crossAx=100)
chart2.x_axis.number_format = 'mm/dd/yyyy'
chart2.x_axis.majorTimeUnit = "months"
values2 = Reference(sheet2, min_col = 2, min_row = 2, max_row = len(incomeData)+1) 
series2 = Series(values2, title="Dividends")
chart2.append(series2)
dates2 = Reference(sheet2, min_col=1, min_row=2, max_row = len(incomeData)+1)
chart2.set_categories(dates2) 
sheet.add_chart(chart2, "G26") 
#sheet2.add_chart(chart2, "E2")

chart3 = BarChart()  
chart3.height = 12
chart3.width = 22
chart3.title = " Annual Returns "
chart3.x_axis.title = " Date "
chart3.y_axis.title = " Percent "
chart3.y_axis.crossAx = 500
chart3.x_axis = DateAxis(crossAx=100)
chart3.x_axis.number_format = 'mm/dd/yyyy'
chart3.x_axis.majorTimeUnit = "years"
values3 = Reference(sheet3, min_col = 2, min_row = 2, max_row = len(annualReturnsData)+1) 
series3 = Series(values3, title="Annual Returns")
chart3.append(series3)
dates3 = Reference(sheet3, min_col=1, min_row=2, max_row = len(annualReturnsData)+1)
chart3.set_categories(dates3) 
sheet.add_chart(chart3, "G50") 
#sheet3.add_chart(chart2, "E2")

writer.save()
writer.close()

'''
cols = [0,1]
data20 = pd.read_excel('./20-Year-Treasury-Constant-Maturity-Rate.xlsx',sheet_name=0,usecols=cols)
data30 = pd.read_excel('./30-Year-Treasury-Constant-Maturity-Rate.xlsx',sheet_name=0,usecols=cols)
pDates = [(2, 15), (5,15), (8,15), (11, 15)]
startDate = datetime.date(2002, 7, 1)
TLT = BondETF(startDate, data20, data30, pDates, 81.75)
idx = bisect.bisect_left(data30.Date, startDate)
marketRate = data30.Yield[idx]
print(TLT.getMarketValue(marketRate, startDate))
assets = TLT.assets.getArray()
for asset in assets:
    print("\n\n" + str(asset))
'''
'''
cols = [0,1]
data20 = pd.read_excel('./20-Year-Treasury-Constant-Maturity-Rate.xlsx',sheet_name=0,usecols=cols)
data30 = pd.read_excel('./30-Year-Treasury-Constant-Maturity-Rate.xlsx',sheet_name=0,usecols=cols)
treasuryData20 = [(data20.Date[i].date(),data20.Yield[i]) for i in range(3770, len(data20.Date)) if (data20.Date[i].date().month - data20.Date[i-1].date().month)%12 == 1]
treasuryData30 = [(data30.Date[i].date(),data30.Yield[i]) for i in range(1, len(data30.Date)) if (data30.Date[i].date().month - data30.Date[i-1].date().month)%12 == 1]
#for i in range(len(treasuryData30)):
    #print("20 Year  " + treasuryData20[i][0].strftime("%m/%d/%y") + "   " + str(treasuryData20[i][1]))
    #print("30 Year  " + treasuryData30[i][0].strftime("%m/%d/%y") + "   " + str(treasuryData30[i][1]))

tltAdjusted = []
tltAdjusted.append((treasuryData30[0][0],1000))
for i in range(1,len(treasuryData30)):
    previousPrice = tltAdjusted[len(tltAdjusted)-1][1]
    currentPrice = 0
    sum = 0
    for j in np.arange(0.0, 10.0, 0.5):
        #linear interpolation
        previousRate = (treasuryData20[i-1][1]/100.) + ((10-j)*((treasuryData30[i-1][1]/100.)-(treasuryData20[i-1][1]/100.))/(10.))
        currentRate = (treasuryData20[i][1]/100.) + ((10-j)*((treasuryData30[i][1]/100.)-(treasuryData20[i][1])/100.)/(10.))
        
        #average
        #previousRate = ((treasuryData20[i-1][1]/100.)+(treasuryData30[i-1][1]/100.))/2.
        #currentRate = ((treasuryData20[i][1]/100.)+(treasuryData30[i][1]/100.))/2.
        #(1-(30-j)/10)20year + ((30-j)/10)30year
        
        #Rounded
        #if 30-j >= 25:
            #previousRate = treasuryData30[i-1][1]/100.
            #currentRate = treasuryData30[i][1]/100.
        #else:
            #previousRate = treasuryData20[i-1][1]/100.
            #currentRate = treasuryData20[i][1]/100.
        
        #factor = pow(1+currentRate,-1*(30-j))
        #sum += previousPrice*((previousRate)*(1-factor)/(currentRate) + factor) #+ ((previousRate+currentRate)/2)/12)
        factor = pow(1+currentRate/2,-1*2*(30-j))
        sum += previousPrice*((previousRate/2)*(1-factor)/(currentRate/2) + factor) #+ ((previousRate+currentRate)/4)/12)
        #currentPrice = previousPrice*(((((previousRate*(1-pow((1+currentRate),(30*-1)))/currentRate) + (1/pow((1+currentRate),30))-1)) + (((previousRate+currentRate)/2)/12))+1)
        currentPrice = sum/20.
    tltAdjusted.append((treasuryData30[i][0],currentPrice))

fileName = './tlt.xlsx'
sheetName = 'TLT4'
book = load_workbook(fileName)
writer = pd.ExcelWriter(fileName, engine = 'openpyxl')
df = pd.DataFrame(tltAdjusted, columns=['Date', 'Value'])
try:
    book.remove(book[sheetName])
except:
    print("worksheet doesn't exist")
finally:
    writer.book = book
    df.to_excel(writer, sheet_name = sheetName, index=False)

wb = writer.book
sheet = wb[sheetName]
#sheet2 = wb['TLT']
chart = LineChart()  
chart.height = 12
chart.width = 22
chart.title = " TLT "
chart.x_axis.title = " Date "
chart.y_axis.title = " Price "
chart.y_axis.crossAx = 500
chart.x_axis = DateAxis(crossAx=100)
chart.x_axis.number_format = 'mm/dd/yyyy'
chart.x_axis.majorTimeUnit = "months"
values1 = Reference(sheet, min_col = 2, min_row = 2, max_row = len(tltAdjusted)+1) 
series1 = Series(values1, title="Calculated TLT")
chart.append(series1)
#values2 = Reference(sheet2, min_col = 2, min_row = 2, max_row = 222) 
#series2 = Series(values2, title="Actual TLT")
#chart.append(series2)
dates = Reference(sheet, min_col=1, min_row=2, max_row = len(tltAdjusted)+1)
chart.set_categories(dates) 
sheet.add_chart(chart, "E2") 


writer.save()
writer.close()
'''
'''
#TLT 3
cols = [0,1]
data20 = pd.read_excel('./20-Year-Treasury-Constant-Maturity-Rate.xlsx',sheet_name=0,usecols=cols)
data30 = pd.read_excel('./30-Year-Treasury-Constant-Maturity-Rate.xlsx',sheet_name=0,usecols=cols)
treasuryData20 = [(data20.Date[i].date(),data20.Yield[i]) for i in range(10107, len(data20.Date)) if (data20.Date[i].date().month - data20.Date[i-1].date().month)%12 == 1]
treasuryData30 = [(data30.Date[i].date(),data30.Yield[i]) for i in range(6335, len(data30.Date)) if (data30.Date[i].date().month - data30.Date[i-1].date().month)%12 == 1]
#for i in range(len(treasuryData30)):
    #print("20 Year  " + treasuryData20[i][0].strftime("%m/%d/%y") + "   " + str(treasuryData20[i][1]))
    #print("30 Year  " + treasuryData30[i][0].strftime("%m/%d/%y") + "   " + str(treasuryData30[i][1]))

tltAdjusted = []
tltAdjusted.append((treasuryData30[0][0],81.75))
for i in range(1,len(treasuryData30)):
    previousPrice = tltAdjusted[len(tltAdjusted)-1][1]
    currentPrice = 0
    sum = 0
    for j in np.arange(0.0, 10.5, 0.5):
        #Most accurate
        #linear interpolation
        previousRate = (treasuryData20[i-1][1]/100.) + ((10-j)*((treasuryData30[i-1][1]/100.)-(treasuryData20[i-1][1]/100.))/(10.))
        currentRate = (treasuryData20[i][1]/100.) + ((10-j)*((treasuryData30[i][1]/100.)-(treasuryData20[i][1])/100.)/(10.))
        
        #average
        #previousRate = ((treasuryData20[i-1][1]/100.)+(treasuryData30[i-1][1]/100.))/2.
        #currentRate = ((treasuryData20[i][1]/100.)+(treasuryData30[i][1]/100.))/2.
        #(1-(30-j)/10)20year + ((30-j)/10)30year
       
        #Rounding to nearest maturity
        #if 30-j >= 25:
            #previousRate = treasuryData30[i-1][1]/100.
            #currentRate = treasuryData30[i][1]/100.
        #else:
            #previousRate = treasuryData20[i-1][1]/100.
            #currentRate = treasuryData20[i][1]/100.     

        #Using only 30 year
        #previousRate = treasuryData30[i-1][1]/100.
        #currentRate = treasuryData30[i][1]/100.

        #factor = pow(1+currentRate,-1*(30-j))
        #sum += previousPrice*((previousRate)*(1-factor)/(currentRate) + factor) #+ ((previousRate+currentRate)/2)/12)
        factor = pow(1+currentRate/2,-1*2*(30-j))
        sum += previousPrice*((previousRate/2)*(1-factor)/(currentRate/2) + factor) #+ ((previousRate+currentRate)/4)/12)
        #currentPrice = previousPrice*(((((previousRate*(1-pow((1+currentRate),(30*-1)))/currentRate) + (1/pow((1+currentRate),30))-1)) + (((previousRate+currentRate)/2)/12))+1)
    currentPrice = sum/21.
    tltAdjusted.append((treasuryData30[i][0],currentPrice))

fileName = './tlt.xlsx'
sheetName = 'TLT3'
book = load_workbook(fileName)
writer = pd.ExcelWriter(fileName, engine = 'openpyxl')
df = pd.DataFrame(tltAdjusted, columns=['Date', 'Value'])
try:
    book.remove(book[sheetName])
except:
    print("worksheet doesn't exist")
finally:
    writer.book = book
    df.to_excel(writer, sheet_name = sheetName, index=False)

wb = writer.book
sheet = wb[sheetName]
sheet2 = wb['TLT']
chart = LineChart()  
chart.height = 12
chart.width = 22
chart.title = " TLT "
chart.x_axis.title = " Date "
chart.y_axis.title = " Price "
chart.y_axis.crossAx = 500
chart.x_axis = DateAxis(crossAx=100)
chart.x_axis.number_format = 'mm/dd/yyyy'
chart.x_axis.majorTimeUnit = "months"
values1 = Reference(sheet, min_col = 2, min_row = 2, max_row = len(tltAdjusted)+1) 
series1 = Series(values1, title="Calculated TLT")
chart.append(series1)
values2 = Reference(sheet2, min_col = 2, min_row = 2, max_row = 222) 
series2 = Series(values2, title="Actual TLT")
chart.append(series2)
dates = Reference(sheet, min_col=1, min_row=2, max_row = len(tltAdjusted)+1)
chart.set_categories(dates) 
sheet.add_chart(chart, "E2") 


writer.save()
writer.close()
'''
'''
#Obtaining Yield data
cols = [0,1]
data20 = pd.read_excel('./20-Year-Treasury-Constant-Maturity-Rate.xlsx',sheet_name=0,usecols=cols)
data30 = pd.read_excel('./30-Year-Treasury-Constant-Maturity-Rate.xlsx',sheet_name=0,usecols=cols)
pre1977 = [(data20.Date[i].date(),data20.Yield[i]) for i in range(2002, 3770) if (data20.Date[i+1].date().month - data20.Date[i].date().month)%12 == 1]
treasuryData20 = pre1977 + [(data20.Date[i].date(),data20.Yield[i]) for i in range(3770, len(data20.Date)-1) if (data20.Date[i+1].date().month - data20.Date[i].date().month)%12 == 1]
treasuryData30 = pre1977 +  [(data30.Date[i].date(),data30.Yield[i]) for i in range(1, len(data30.Date)-1) if (data30.Date[i+1].date().month - data30.Date[i].date().month)%12 == 1]

#print("# of 20 year dates: " + str(len(treasuryData20)))
#print("# of 20 year dates: " + str(len(treasuryData20)))
#for i in range(len(treasuryData30)):
    #print("20 Year  " + treasuryData20[i][0].strftime("%m/%d/%y") + "   " + str(treasuryData20[i][1]))
    #print("30 Year  " + treasuryData30[i][0].strftime("%m/%d/%y") + "   " + str(treasuryData30[i][1]))

tltAdjusted = []
tltAdjusted.append((treasuryData30[0][0],1000, 0))
for i in range(1,len(treasuryData30)):
    previousPrice = tltAdjusted[len(tltAdjusted)-1][1]
    currentPrice = 0
    coupon = 0
    sum = 0
    for j in np.arange(0.0, 10.5, 0.5):
        #linear interpolation
        previousRate = (treasuryData20[i-1][1]/100.) + ((10-j)*((treasuryData30[i-1][1]/100.)-(treasuryData20[i-1][1]/100.))/(10.))
        currentRate = (treasuryData20[i][1]/100.) + ((10-j)*((treasuryData30[i][1]/100.)-(treasuryData20[i][1])/100.)/(10.))
        
        #average
        #previousRate = ((treasuryData20[i-1][1]/100.)+(treasuryData30[i-1][1]/100.))/2.
        #currentRate = ((treasuryData20[i][1]/100.)+(treasuryData30[i][1]/100.))/2.
        #(1-(30-j)/10)20year + ((30-j)/10)30year
        
        #Rounded
        #if 30-j >= 25:
            #previousRate = treasuryData30[i-1][1]/100.
            #currentRate = treasuryData30[i][1]/100.
        #else:
            #previousRate = treasuryData20[i-1][1]/100.
            #currentRate = treasuryData20[i][1]/100.
        
        #factor = pow(1+currentRate,-1*(30-j))
        #sum += previousPrice*((previousRate)*(1-factor)/(currentRate) + factor) #+ ((previousRate+currentRate)/2)/12)
        factor = pow(1+currentRate/2,-1*2*(30-j))
        yld = 2*((previousRate/2.)*(currentRate/2.))/((previousRate/2.)+factor*((currentRate/2.)-(previousRate/2.)))
        presentValue = (previousPrice*((previousRate/2)*(1-factor)/(currentRate/2) + factor))/21. #+ ((previousRate+currentRate)/4)/12)
        coupon += yld*presentValue
        sum += presentValue
        #currentPrice = previousPrice*(((((previousRate*(1-pow((1+currentRate),(30*-1)))/currentRate) + (1/pow((1+currentRate),30))-1)) + (((previousRate+currentRate)/2)/12))+1)
    currentPrice = sum
    divYld = coupon/sum
    tltAdjusted.append((treasuryData30[i][0],currentPrice,divYld*100.))

fileName = './tlt.xlsx'
sheetName = 'TLT Price and Yield (1970)'
book = load_workbook(fileName)
writer = pd.ExcelWriter(fileName, engine = 'openpyxl')
df = pd.DataFrame(tltAdjusted, columns=['Date', 'Value', 'Yield'])
try:
    book.remove(book[sheetName])
except:
    print("worksheet doesn't exist")
finally:
    writer.book = book
    df.to_excel(writer, sheet_name = sheetName, index=False)

wb = writer.book
sheet = wb[sheetName]
#sheet2 = wb['TLT']
chart = LineChart()  
chart.height = 12
chart.width = 22
chart.title = " TLT "
chart.x_axis.title = " Date "
chart.y_axis.title = " Price "
chart.y_axis.crossAx = 500
chart.x_axis = DateAxis(crossAx=100)
chart.x_axis.number_format = 'mm/dd/yyyy'
chart.x_axis.majorTimeUnit = "months"
values1 = Reference(sheet, min_col = 2, min_row = 2, max_row = len(tltAdjusted)+1) 
series1 = Series(values1, title="Calculated TLT")
chart.append(series1)
dates = Reference(sheet, min_col=1, min_row=2, max_row = len(tltAdjusted)+1)
chart.set_categories(dates) 
sheet.add_chart(chart, "E2") 

chart2 = LineChart()  
chart2.height = 12
chart2.width = 22
chart2.title = " TLT Yield "
chart2.x_axis.title = " Date "
chart2.y_axis.title = " Yield "
chart2.y_axis.crossAx = 500
chart2.x_axis = DateAxis(crossAx=100)
chart2.x_axis.number_format = 'mm/dd/yyyy'
chart2.x_axis.majorTimeUnit = "months"
values2 = Reference(sheet, min_col = 3, min_row = 2, max_row = len(tltAdjusted)+1) 
series2 = Series(values2, title="Yield")
chart2.append(series2)
chart2.set_categories(dates) 
sheet.add_chart(chart2, "E26") 


writer.save()
writer.close()
'''

'''
#Total return bonds
cols = [0,1,2]
dat = pd.read_excel('./tlt.xlsx',sheet_name="TLT Price and Yield (1970)",usecols=cols)
priceData = [(dat.Date[i].date(),dat.Value[i]) for i in range(len(dat.Date))]
yieldData = [(dat.Date[i].date(),dat.Yield[i]) for i in range(len(dat.Date))]
#for i in range(len(treasuryData30)):
    #print("20 Year  " + treasuryData20[i][0].strftime("%m/%d/%y") + "   " + str(treasuryData20[i][1]))
    #print("30 Year  " + treasuryData30[i][0].strftime("%m/%d/%y") + "   " + str(treasuryData30[i][1]))

tltAdjusted = []
tltAdjusted.append((priceData[0][0],10000))
for i in range(1,len(priceData)):
    previousPrice = tltAdjusted[len(tltAdjusted)-1][1]
    yld = yieldData[i][1]
    growth = (priceData[i][1]-priceData[i-1][1])/priceData[i-1][1]
    currentPrice = previousPrice*(1+growth)
    currentPrice += currentPrice*((yld/100.)/12)
    tltAdjusted.append((priceData[i][0], currentPrice))

fileName = './tlt.xlsx'
sheetName = 'TLT Total Return (1970)'
book = load_workbook(fileName)
writer = pd.ExcelWriter(fileName, engine = 'openpyxl')
df = pd.DataFrame(tltAdjusted, columns=['Date', 'Value'])
try:
    book.remove(book[sheetName])
except:
    print("worksheet doesn't exist")
finally:
    writer.book = book
    df.to_excel(writer, sheet_name = sheetName, index=False)

wb = writer.book
sheet = wb[sheetName]
chart = LineChart()  
chart.height = 12
chart.width = 22
chart.title = " Total Return TLT "
chart.x_axis.title = " Date "
chart.y_axis.title = " Price "
chart.y_axis.crossAx = 500
chart.x_axis = DateAxis(crossAx=100)
chart.x_axis.number_format = 'mm/dd/yyyy'
chart.x_axis.majorTimeUnit = "months"
values1 = Reference(sheet, min_col = 2, min_row = 2, max_row = len(tltAdjusted)+1) 
series1 = Series(values1, title="Calculated TLT")
chart.append(series1)
dates = Reference(sheet, min_col=1, min_row=2, max_row = len(tltAdjusted)+1)
chart.set_categories(dates) 
sheet.add_chart(chart, "E2") 

writer.save()
writer.close()
'''





'''
df = pd.DataFrame(tltAdjusted, columns=['Date', 'Value', 'Yield'])
df.to_excel('./tlt.xlsx', sheet_name='S&P500_TOTAL_RETURN', index=False)

#p(1+r)^t=A this is the formula for t-bills; p is price; r is interest rate; t is time for maturity in years; a is the amount of money the government will pay you when the bill matures 

x = datetime.date(2016, 2, 15)
y = datetime.date(2020, 7, 31)
b = Bond(1085405500,2.5,x,30)
for r in range(980,1201):
    rf = r/1000.
    #if p%10000 == 195552:
    print("Clean value: " + str(b.getCleanPrice(rf,y)) )
    print("Market value: " + str(b.presentValue(rf,y)) + "   @ Market rate: " + str(rf) + "\n\n")

spdata = pd.read_excel('./tmp.xlsx',usecols=cols)
dates = spdata.Date
close = spdata.Close
data = [(dates[i],close[i]) for i in range(len(dates))]

print(data)


cols = [0,4]
spPrice = pd.read_excel('./S&P500.xlsx',sheet_name=0,usecols=cols)
#dates = spPrice.Date
#close = spPrice.Close
spPriceTuple = [(spPrice.Date[i].date(),spPrice.Close[i]) for i in range(len(spPrice.Date))]
#print(spPriceTuple)

cols = [0,1]
spYield = pd.read_excel('./S&P500.xlsx',sheet_name=3,usecols=cols)
#dates = spYield.Date
#close = spYield.Value
spYieldTuple = [(spYield.Date[i].date(),spYield.Value[i]) for i in range(683, len(spYield.Date),3)]

spTotalReturn = []
quarterIter = 0
for i in range(len(spPriceTuple)):
     if i == 0:
         spTotalReturn.append((spPriceTuple[i][0],17.66, 0, 0))
     else:
         dailyVal = spTotalReturn[len(spTotalReturn)-1][1]
         yld = 0
         div = 0
         m1 = spPriceTuple[i][0].month
         m2 = spPriceTuple[i-1][0].month
         if (m1 == 1 and m2 == 12) or (m1 == 4 and m2 == 3) or (m1 == 7 and m2 == 6 ) or (m1 == 10 and m2 == 9):
             yld = spYieldTuple[quarterIter][1]
             div = dailyVal*((yld/100.)/4)
             dailyVal += div
             quarterIter += 1
         growth = (spPriceTuple[i][1]-spPriceTuple[i-1][1])/spPriceTuple[i-1][1]
         dailyVal *= (1 + growth)
         spTotalReturn.append((spPriceTuple[i][0],dailyVal,yld,div))


for (d, v, y, div) in spTotalReturn:
    if div == 0:
        print(d.strftime("%m/%d/%y") + "   Value: " + str(v))
    else: 
        print(d.strftime("%m/%d/%y") + "   Value: " + str(v) + "   Yield: " + str(y) + "   dividend received: " + str(div))


df = pd.DataFrame(spTotalReturn, columns=['Date', 'Value', 'Yield', 'Dividend Received'])
df.to_excel('./S&P500.xlsx', sheet_name='S&P500_TOTAL_RETURN', index=False)


x = datetime.date(1977, 2, 15)
y = datetime.date(1987, 2, 17)
b = Bond(1000,7.7,x,30)
presentValue = round(b.getNumberOfRemainingPayments(y),2)
print("Number of remaining payments: " + str(presentValue))
print("Value of bond on " + y.strftime("%m/%d/%y") + ": " + str(b.presentValue(7.48, y)))
day_count = (b.getMaturityDate()-x).days + 1
numDays = 0
for single_date in [d for d in (x + timedelta(n) for n in range(day_count)) if d <= b.getMaturityDate()]:
    temp = b.getNumberOfRemainingPayments(single_date)
    if temp != numDays:
        numDays = temp
        print(single_date.strftime("%m/%d/%y") + ": number of payments remaining - " + str(numDays))
day_count = (b.getMaturityDate()-x).days + 1
total = 0
for single_date in [d for d in (x + timedelta(n) for n in range(day_count)) if d <= b.getMaturityDate()]:
    payment = b.getPayment(single_date)
    if payment > 0:
        total += payment
        print(payment)
print("total return: " + str(total))
'''

